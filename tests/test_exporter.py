import os
import pytest
import openpyxl
from PyQt6.QtWidgets import QApplication  # ДОДАНО
from app.services.exporter import Exporter

# ДОДАНО: Ініціалізуємо додаток Qt для рендерингу PDF (без цього QTextDocument падає)
app = QApplication.instance() or QApplication([])


# ─── ТЕСТ 1: Генерація Excel при 0 записах ────────────────────────────────

def test_export_excel_empty(tmp_path):
    filepath = str(tmp_path / "empty_report.xlsx")
    headers = ["Інв. номер", "Бренд", "Модель"]
    data = []

    Exporter.to_excel(filepath, headers, data, "Порожній звіт")

    assert os.path.exists(filepath)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "Інв. номер"


# ─── ТЕСТ 2: Генерація Excel при 100+ записах ─────────────────────────────

def test_export_excel_bulk(tmp_path):
    filepath = str(tmp_path / "bulk_report.xlsx")
    headers = ["ID", "Назва", "Значення"]

    data = [[i, f"Елемент {i}", i * 10] for i in range(150)]

    Exporter.to_excel(filepath, headers, data, "Великий звіт")

    assert os.path.exists(filepath)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    assert ws.max_row == 152
    assert ws.cell(row=152, column=2).value == "Елемент 149"


# ─── ТЕСТ 3: Генерація PDF (базова перевірка) ─────────────────────────────

def test_export_pdf_creation(tmp_path):
    filepath = str(tmp_path / "report.pdf")
    headers = ["Колонка 1", "Колонка 2"]
    data = [["Тест 1", "Тест 2"], ["Дані 1", "Дані 2"]]

    Exporter.to_pdf(filepath, headers, data, "PDF Звіт")

    assert os.path.exists(filepath)
    assert os.path.getsize(filepath) > 0