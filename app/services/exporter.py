import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter  # ДОДАНО: безпечне отримання літери колонки
from PyQt6.QtGui import QTextDocument, QPdfWriter, QPageSize, QPageLayout
from PyQt6.QtCore import QMarginsF


class Exporter:
    @staticmethod
    def to_excel(filepath: str, headers: list[str], data: list[list], title: str = "Звіт"):
        """Експорт у формат Excel (.xlsx) з базовим форматуванням."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Звіт"

        # Заголовок звіту
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws['A1']
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Заголовки колонок
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col).font = Font(bold=True)

        # Дані
        for row in data:
            ws.append(row)

        # Автоматичне розширення колонок
        # enumerate(..., 1) дає нам номер колонки починаючи з 1
        for idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(idx)  # Безпечно отримуємо літеру (A, B, C...)
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filepath)

    @staticmethod
    def to_pdf(filepath: str, headers: list[str], data: list[list], title: str = "Звіт"):
        """Експорт у PDF за допомогою вбудованого рушія Qt (ідеально для кирилиці)."""
        # Формуємо HTML таблицю
        html = f"<h1 style='text-align: center; font-family: sans-serif;'>{title}</h1>"
        html += "<table border='1' cellspacing='0' cellpadding='5' width='100%' style='font-family: sans-serif; font-size: 10px; border-collapse: collapse;'>"
        html += "<tr style='background-color: #f2f2f2;'>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"

        for row in data:
            html += "<tr>" + "".join(f"<td>{str(cell) if cell is not None else ''}</td>" for cell in row) + "</tr>"
        html += "</table>"

        doc = QTextDocument()
        doc.setHtml(html)

        # Налаштовуємо PDF Writer
        writer = QPdfWriter(filepath)
        writer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        writer.setPageOrientation(QPageLayout.Orientation.Landscape)

        # Відступи (margins)
        layout = writer.pageLayout()
        layout.setMargins(QMarginsF(10, 10, 10, 10))
        writer.setPageLayout(layout)

        doc.print(writer)